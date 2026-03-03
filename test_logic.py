"""
Unit tests for Stripe Reconciliation Accelerator
"""

import unittest
from io import BytesIO
import pandas as pd
from logic import StripeReconciliator


class TestStripeReconciliator(unittest.TestCase):
    """Test suite for StripeReconciliator class."""

    def setUp(self):
        """Set up test fixtures."""
        self.reconciliator = StripeReconciliator()
        
        # Create sample CSV data
        self.balance_csv = BytesIO(b"""Date,Payout ID,Charge Amount,Refund Amount,Fee Amount,Dispute Amount,Balance
2024-02-28,payout_001,5000.00,-200.00,-120.50,0.00,4679.50
2024-02-27,payout_002,3500.00,-150.00,-85.25,-50.00,3214.75""")
        
        self.payout_csv = BytesIO(b"""Payout ID,Payout Date,Payout Amount,Currency,Status,Destination,Arrival Date
payout_001,2024-02-27,4679.50,USD,paid,Bank Account ****3210,2024-02-28
payout_002,2024-02-26,3214.75,USD,paid,Bank Account ****3210,2024-02-27""")
        
        self.bank_csv = BytesIO(b"""Date,Description,Debit,Credit,Balance,Reference
2024-02-28,STRIPE TRANSFER 12345,0.00,4679.50,125678.50,REF001
2024-02-27,STRIPE TRANSFER 12346,0.00,3214.75,125000.00,REF002""")

    def test_load_csv(self):
        """Test CSV loading functionality."""
        self.balance_csv.seek(0)
        result = self.reconciliator.load_csv(self.balance_csv.getvalue(), 'balance')
        self.assertTrue(result)
        self.assertIsNotNone(self.reconciliator.balance_df)
        self.assertEqual(len(self.reconciliator.balance_df), 2)

    def test_clean_data(self):
        """Test data cleaning."""
        self.balance_csv.seek(0)
        self.reconciliator.load_csv(self.balance_csv.getvalue(), 'balance')
        self.reconciliator.clean_data()
        
        # Verify whitespace stripping
        self.assertIsNotNone(self.reconciliator.balance_df)

    def test_calculate_net_activity(self):
        """Test net activity calculation."""
        self.balance_csv.seek(0)
        self.reconciliator.load_csv(self.balance_csv.getvalue(), 'balance')
        
        activity = self.reconciliator.calculate_net_activity()
        self.assertIsInstance(activity, dict)
        self.assertTrue('charges' in activity)

    def test_run_reconciliation(self):
        """Verify the new matching engine and summary logic."""
        self.balance_csv.seek(0)
        self.payout_csv.seek(0)
        self.bank_csv.seek(0)

        self.reconciliator.load_csv(self.balance_csv.getvalue(), 'balance')
        self.reconciliator.load_csv(self.payout_csv.getvalue(), 'payout')
        self.reconciliator.load_csv(self.bank_csv.getvalue(), 'bank')

        matched, exceptions, summary = self.reconciliator.run_reconciliation(tolerance=5.0)
        # both payouts should be matched exactly
        self.assertEqual(len(matched), 2)
        self.assertTrue(all(matched['match_type'] == 'MATCHED'))
        self.assertEqual(len(exceptions), 0)
        self.assertEqual(summary['status'], 'RECONCILED')

    def test_report_generation(self):
        """Test Excel report generation and verify sheets."""
        self.balance_csv.seek(0)
        self.payout_csv.seek(0)
        self.bank_csv.seek(0)

        self.reconciliator.load_csv(self.balance_csv.getvalue(), 'balance')
        self.reconciliator.load_csv(self.payout_csv.getvalue(), 'payout')
        self.reconciliator.load_csv(self.bank_csv.getvalue(), 'bank')

        report = self.reconciliator.generate_report()
        self.assertIsNotNone(report)
        self.assertIsInstance(report, BytesIO)
        # ensure report has expected sheets
        from openpyxl import load_workbook
        report.seek(0)
        wb = load_workbook(report)
        self.assertIn('Executive_Summary', wb.sheetnames)
        self.assertIn('Payout_Matching', wb.sheetnames)
        self.assertIn('Exceptions', wb.sheetnames)
        self.assertIn('Audit_Log', wb.sheetnames)


if __name__ == '__main__':
    unittest.main()
