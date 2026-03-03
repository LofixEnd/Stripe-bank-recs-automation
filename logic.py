"""
Stripe Payout Reconciliation Automation Tool
Production-ready script for monthly Stripe reconciliation in eCommerce bookkeeping
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO, StringIO
import logging
import sys
from pathlib import Path
from typing import Tuple, Dict, List
import os
import difflib


# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================
# Configure logging to track file operations, calculations, and exceptions
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('stripe_reconciliation.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# ============================================================================
# DATA LOADING AND VALIDATION FUNCTIONS
# ============================================================================

def get_column_mapping(
    actual_columns: List[str],
    column_variations: Dict[str, List[str]],
    required_columns: List[str],
    optional_columns: Dict[str, str] = None
) -> Dict[str, str]:
    """
    Flexible column mapping function that handles variations in column names.
    
    Args:
        actual_columns: List of column names from the CSV file (normalized)
        column_variations: Dict mapping standard name -> list of possible variations
        required_columns: List of column names that must be present
        optional_columns: Dict mapping standard name -> default value for missing optional columns
    
    Returns:
        Dict mapping standard column name -> actual column name in CSV
    
    Raises:
        ValueError: If critical required columns cannot be mapped
    """
    if optional_columns is None:
        optional_columns = {}
    
    mapping = {}
    unmapped_required = []

    # normalize actual columns for matching
    actual_norm_map = {normalize_name(col): col for col in actual_columns}
    actual_norm_keys = list(actual_norm_map.keys())

    # Try to map each standard column to an actual column
    for standard_col in required_columns + list(optional_columns.keys()):
        variations = column_variations.get(standard_col, [])
        # include the standard_col itself as a variation
        variations.append(standard_col)
        # normalize variations
        variations_norm = [normalize_name(v) for v in variations]

        found = False
        matched_actual = None

        # exact normalized match
        for var_norm in variations_norm:
            if var_norm in actual_norm_map:
                matched_actual = actual_norm_map[var_norm]
                found = True
                break

        # fuzzy match fallback
        if not found:
            for var_norm in variations_norm:
                close = difflib.get_close_matches(var_norm, actual_norm_keys, n=1, cutoff=0.8)
                if close:
                    matched_actual = actual_norm_map[close[0]]
                    found = True
                    logger.info(f"Fuzzy-matched column '{standard_col}' to '{matched_actual}' using variation '{var_norm}'")
                    break

        if found and matched_actual:
            mapping[standard_col] = matched_actual
        else:
            if standard_col in required_columns:
                unmapped_required.append(standard_col)
                logger.error(
                    f"Critical column '{standard_col}' not found. "
                    f"Looked for variations: {variations}. "
                    f"Available columns: {actual_columns}"
                )
            elif standard_col in optional_columns:
                logger.warning(
                    f"Optional column '{standard_col}' not found (variations: {variations}). "
                    f"Will use default value: {optional_columns[standard_col]}"
                )

    if unmapped_required:
        raise ValueError(
            f"Cannot map required columns: {unmapped_required}. "
            f"Available columns: {actual_columns}"
        )

    # log any actual columns that were not mapped at all
    unused = [col for col in actual_columns if col not in mapping.values()]
    if unused:
        logger.info(f"Unused columns in file: {unused}")

    return mapping


def rename_columns_based_on_mapping(df: pd.DataFrame, mapping: Dict[str, str]) -> pd.DataFrame:
    """
    Rename DataFrame columns based on mapping, and add default values for missing optional columns.
    
    Args:
        df: DataFrame with original column names
        mapping: Dict mapping standard name -> actual column name
    
    Returns:
        DataFrame with standardized column names
    """
    # Create reverse mapping (actual -> standard)
    rename_dict = {v: k for k, v in mapping.items()}
    
    # Rename columns
    df = df.rename(columns=rename_dict)
    
    return df


def normalize_name(name: str) -> str:
    """Normalize column name for comparison: lowercase, strip non-alphanumerics."""
    return ''.join(ch for ch in name.lower() if ch.isalnum())


def load_data_file(file_path: str) -> pd.DataFrame:
    """
    Load a data file (CSV or Excel) with error handling and normalization.
    Detects format by extension, handles encoding issues, and strips
    whitespace from headers and string values.
    """
    try:
        if not Path(file_path).exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = Path(file_path).suffix.lower()
        if ext in ['.csv']:
            try:
                df = pd.read_csv(file_path, encoding='utf-8')
            except UnicodeDecodeError:
                logger.warning(f"UTF-8 encoding failed for {file_path}, trying latin-1...")
                df = pd.read_csv(file_path, encoding='latin-1')
        elif ext in ['.xls', '.xlsx']:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type {ext} for {file_path}")

        logger.info(f"Successfully loaded file: {file_path}")

        if len(df) == 0:
            raise ValueError(f"CSV/Excel file is empty, cannot continue: {file_path}")

        # Strip whitespace from column names
        df.columns = df.columns.str.strip()

        # Strip whitespace from string columns
        for col in df.select_dtypes(include=['object', 'string']).columns:
            df[col] = df[col].astype(str).str.strip()

        logger.info(f"Loaded {len(df)} rows from {file_path}")
        return df

    except FileNotFoundError as e:
        logger.error(str(e))
        raise
    except ValueError as e:
        logger.error(f"Validation error in {file_path}: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Error loading {file_path}: {str(e)}")
        raise


def load_stripe_balance_report(file_path: str) -> pd.DataFrame:
    """
    Load Stripe Balance Report with flexible column mapping.
    
    Maps possible variations of column names:
    - created: created, transaction_date
    - available_on: available_on, available_date, created (fallback)
    - type: type, transaction_type, description (fallback)
    - amount: amount, gross, transaction_amount
    - fee: fee, transaction_fee, fees
    - net: net, net_amount, amount (fallback if amount not available)
    - payout_id: payout_id, id
    - currency: currency (optional, defaults to USD)
    
    Converts datetime and numeric columns to appropriate types.
    Handles NaN payout_id values.
    """
    # Load file (CSV or Excel) with flexible encoding
    df = load_data_file(file_path)
    
    # Define column name variations for Stripe Balance Report
    column_variations = {
        'created': ['created', 'transaction_date', 'date'],
        'available_on': ['available_on', 'available_date'],
        'type': ['type', 'transaction_type'],
        'amount': ['amount', 'gross', 'transaction_amount'],
        'fee': ['fee', 'transaction_fee', 'fees'],
        'net': ['net', 'net_amount'],
        'payout_id': ['payout_id', 'id', 'payout'],
        'currency': ['currency', 'currency_code']
    }
    
    # Required columns for Stripe Balance Report
    required_columns = ['created', 'type', 'amount', 'fee', 'net', 'payout_id']
    
    # Optional columns with defaults
    optional_columns = {
        'available_on': 'created',  # Use created date if available_on not present
        'currency': 'USD'  # Default to USD if not specified
    }
    
    # Get column mapping
    mapping = get_column_mapping(
        actual_columns=list(df.columns),
        column_variations=column_variations,
        required_columns=required_columns,
        optional_columns=optional_columns
    )
    
    logger.info(f"Column mapping for Stripe Balance Report: {mapping}")
    
    # Rename columns based on mapping
    df = rename_columns_based_on_mapping(df, mapping)
    
    # Add default values for missing optional columns
    if 'currency' not in df.columns:
        df['currency'] = 'USD'
        logger.info("Added default currency value: USD")
    
    # If available_on not present, use created date
    if 'available_on' not in df.columns:
        df['available_on'] = df['created']
        logger.info("Using 'created' column values for 'available_on'")
    
    # Convert datetime columns
    df['created'] = pd.to_datetime(df['created'], errors='coerce')
    df['available_on'] = pd.to_datetime(df['available_on'], errors='coerce')
    
    # Convert numeric columns
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
    df['fee'] = pd.to_numeric(df['fee'], errors='coerce')
    df['net'] = pd.to_numeric(df['net'], errors='coerce')
    
    # Handle payout_id - keep as string but mark NaN explicitly
    df['payout_id'] = df['payout_id'].replace('nan', np.nan)
    
    # Remove rows with critical NaN values
    initial_rows = len(df)
    df = df.dropna(subset=['amount', 'fee', 'net'])
    if len(df) < initial_rows:
        logger.warning(f"Removed {initial_rows - len(df)} rows with missing numeric values")
    
    logger.info(f"Processed Stripe Balance Report: {len(df)} valid transactions")
    return df


def load_stripe_payout_report(file_path: str) -> pd.DataFrame:
    """
    Load Stripe Payout Report with flexible column mapping.
    
    Maps possible variations of column names:
    - payout_id: payout_id, id
    - arrival_date: arrival_date, available_on, date, payout_date
    - amount: amount, payout_amount, gross
    - status: status, payout_status
    - currency: currency (optional, defaults to USD)
    
    Converts datetime and numeric columns to appropriate types.
    Converts status to lowercase for case-insensitive matching.
    """
    # Load file (CSV or Excel) with flexible encoding
    df = load_data_file(file_path)
    
    # Define column name variations for Stripe Payout Report
    column_variations = {
        'payout_id': ['payout_id', 'id'],
        'arrival_date': ['arrival_date', 'available_on', 'date', 'payout_date'],
        'amount': ['amount', 'payout_amount', 'gross'],
        'status': ['status', 'payout_status', 'state'],
        'currency': ['currency', 'currency_code']
    }
    
    # Required columns for Stripe Payout Report
    required_columns = ['payout_id', 'arrival_date', 'amount', 'status']
    
    # Optional columns with defaults
    optional_columns = {
        'currency': 'USD'
    }
    
    # Get column mapping
    mapping = get_column_mapping(
        actual_columns=list(df.columns),
        column_variations=column_variations,
        required_columns=required_columns,
        optional_columns=optional_columns
    )
    
    logger.info(f"Column mapping for Stripe Payout Report: {mapping}")
    
    # Rename columns based on mapping
    df = rename_columns_based_on_mapping(df, mapping)
    
    # Add default values for missing optional columns
    if 'currency' not in df.columns:
        df['currency'] = 'USD'
        logger.info("Added default currency value: USD")
    
    # Convert datetime column
    df['arrival_date'] = pd.to_datetime(df['arrival_date'], errors='coerce')
    
    # Convert numeric columns
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
    
    # Convert status to lowercase for case-insensitive matching
    df['status'] = df['status'].str.lower()
    
    logger.info(f"Processed Stripe Payout Report: {len(df)} payouts loaded")
    return df


def load_bank_statement(file_path: str) -> pd.DataFrame:
    """
    Load Bank Statement with flexible column mapping.
    
    Maps possible variations of column names:
    - date: date, posting_date, transaction_date
    - description: description, details, memo
    - amount: amount, credit, debit, deposit
    
    Filters to only include rows where description contains 'stripe' (case-insensitive).
    Converts date and amount columns to appropriate types.
    """
    # Load file (CSV or Excel) with flexible encoding
    df = load_data_file(file_path)
    
    # Define column name variations for Bank Statement
    column_variations = {
        'date': ['date', 'posting_date', 'transaction_date', 'post_date'],
        'description': ['description', 'details', 'memo', 'transaction_description'],
        'amount': ['amount', 'credit', 'debit', 'deposit', 'transaction_amount']
    }
    
    # Required columns for Bank Statement
    required_columns = ['date', 'description', 'amount']
    
    # No optional columns for bank statement
    optional_columns = {}
    
    # Get column mapping
    mapping = get_column_mapping(
        actual_columns=list(df.columns),
        column_variations=column_variations,
        required_columns=required_columns,
        optional_columns=optional_columns
    )
    
    logger.info(f"Column mapping for Bank Statement: {mapping}")
    
    # Rename columns based on mapping
    df = rename_columns_based_on_mapping(df, mapping)
    
    # Convert datetime column
    df['date'] = pd.to_datetime(df['date'], errors='coerce')
    
    # Convert numeric column
    df['amount'] = pd.to_numeric(df['amount'], errors='coerce')
    
    # Filter for Stripe transactions only (case-insensitive)
    initial_rows = len(df)
    df['description_lower'] = df['description'].str.lower()
    df = df[df['description_lower'].str.contains('stripe', na=False)].copy()
    df = df.drop('description_lower', axis=1)
    
    logger.info(f"Filtered Bank Statement: {len(df)} Stripe transactions (filtered from {initial_rows} total)")
    return df


# ============================================================================
# STRIPE MONTHLY SUMMARY CALCULATION
# ============================================================================

def calculate_stripe_monthly_summary(
    balance_df: pd.DataFrame,
    payout_df: pd.DataFrame,
    bank_df: pd.DataFrame,
    opening_balance: float = 0.0,
    tolerance: float = 5.0
) -> Dict[str, float]:
    """
    Calculate Stripe monthly summary metrics.

    Existing metrics (charges, fees, refunds, disputes, net activity, total payouts,
    ending balance) are preserved.  Additional fields are computed using the
    provided opening balance and the bank transactions, and a reconciliation
    status is determined based on the tolerance.

    Returns dictionary with summary metrics.
    """
    summary: Dict[str, float] = {}

    # New working summary fields
    summary['Opening Balance'] = opening_balance
    summary['Total Stripe Payouts'] = payout_df['amount'].sum()
    summary['Total Bank Deposits'] = bank_df['amount'].sum()
    calculated_ending = opening_balance + summary['Total Bank Deposits']
    expected_ending = opening_balance + summary['Total Stripe Payouts']
    summary['Calculated Ending Balance'] = calculated_ending
    summary['Expected Ending Balance'] = expected_ending
    summary['Ending Difference'] = expected_ending - calculated_ending
    summary['Status'] = (
        'RECONCILED' if abs(summary['Ending Difference']) <= tolerance else 'NOT_RECONCILED'
    )

    # Preserve previous summary calculations
    # Calculate totals by transaction type
    summary['Total Charges'] = balance_df[balance_df['type'].str.lower() == 'charge']['amount'].sum()
    summary['Total Fees'] = balance_df['fee'].sum()
    summary['Total Refunds'] = balance_df[balance_df['type'].str.lower() == 'refund']['amount'].sum()
    summary['Total Disputes'] = balance_df[balance_df['type'].str.lower() == 'dispute']['amount'].sum()

    # Calculate net activity (sum of all net amounts)
    summary['Net Activity'] = balance_df['net'].sum()

    # Calculate total payouts (already captured in Total Stripe Payouts)
    summary['Total Payouts'] = summary['Total Stripe Payouts']

    # Calculate ending balance (original metric, now includes opening balance)
    summary['Ending Balance'] = opening_balance + summary['Net Activity'] - summary['Total Payouts']

    # Determine reconciliation status (placeholder - will be updated after exception detection)
    summary['Reconciliation Status'] = 'Pending'

    logger.info("Stripe Monthly Summary calculated:")
    logger.info(f"  Total Charges: {summary['Total Charges']}")
    logger.info(f"  Total Fees: {summary['Total Fees']}")
    logger.info(f"  Net Activity: {summary['Net Activity']}")
    logger.info(f"  Total Payouts: {summary['Total Payouts']}")
    logger.info(f"  Ending Balance: {summary['Ending Balance']}")

    return summary


# ============================================================================
# PAYOUT MATCHING FUNCTIONS
# ============================================================================

def match_stripe_payouts_to_bank(
    payout_df: pd.DataFrame,
    bank_df: pd.DataFrame,
    tolerance: float = 5.0
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Match Stripe payouts with bank deposits.

    Uses amount tolerance and optional date window.  Handles partial/combined
    payouts and uses description text (payout_id, 'stripe', etc.) as a ranking
    factor.  Returns both matching details and any unmatched bank deposits.

    The function now includes defensive data validation ensuring numeric
    columns are properly typed and providing clear errors if not.  This
    prevents invalid comparisons between datetimes and integers.
    """
    """
    Match Stripe payouts with bank deposits.

    Uses amount tolerance and optional date window.  Handles partial/combined
    payouts and uses description text (payout_id, 'stripe', etc.) as a ranking
    factor.  Returns both matching details and any unmatched bank deposits.
    """
    matching_results: List[Dict[str, any]] = []
    bank_df_copy = bank_df.copy()

    # Defensive typing: ensure amount columns are numeric
    if 'amount' not in payout_df.columns:
        raise ValueError("Payout dataframe missing 'amount' column")
    if 'amount' not in bank_df_copy.columns:
        raise ValueError("Bank dataframe missing 'amount' column")

    # ensure payout amounts numeric and detect unexpected types
    if payout_df['amount'].dtype.kind in ('M','O'):
        # could contain datetime values
        raise ValueError("Payout 'amount' column contains non-numeric values (datetime or objects)")
    payout_df['amount'] = pd.to_numeric(payout_df['amount'], errors='coerce')
    if payout_df['amount'].isna().any():
        logger.warning("Some payout amounts could not be converted to numeric and will be dropped")
        payout_df = payout_df[payout_df['amount'].notna()].copy()

    if bank_df_copy['amount'].dtype.kind in ('M','O'):
        raise ValueError("Bank 'amount' column contains non-numeric values (possibly dates)")
    bank_df_copy['amount'] = pd.to_numeric(bank_df_copy['amount'], errors='coerce')
    if bank_df_copy['amount'].isna().any():
        logger.warning("Some bank deposit amounts could not be converted to numeric and will be dropped")
        bank_df_copy = bank_df_copy[bank_df_copy['amount'].notna()].copy()


    # Ensure bank dates are datetime
    if 'date' in bank_df_copy.columns:
        bank_df_copy['date'] = pd.to_datetime(bank_df_copy['date'], errors='coerce')

    for idx, payout_row in payout_df.iterrows():
        payout_id = payout_row['payout_id']
        payout_amount = payout_row['amount']
        payout_date = payout_row.get('arrival_date', pd.NaT)
        stripe_currency = payout_row.get('currency', None)

        candidates = bank_df_copy[bank_df_copy['matched'] != True].copy()
        candidates['amount_diff'] = candidates['amount'].apply(
            lambda x: round(abs(payout_amount - x), 2)
        )
        candidates = candidates[candidates['amount_diff'] <= tolerance]
        # initialization for matching
        found = False
        bank_idx = None
        bank_desc = ''

        # description scoring helper
        def desc_score(desc: str) -> int:
            if not isinstance(desc, str):
                return 3
            d = desc.lower()
            if payout_id and payout_id.lower() in d:
                return 0
            if 'stripe' in d or 'transfer' in d:
                return 1
            ratio = difflib.SequenceMatcher(None, payout_id.lower() if payout_id else '', d).ratio()
            if ratio > 0.8:
                return 2
            return 3

        if pd.notna(payout_date) and 'date' in candidates.columns:
            # ensure date column is datetime
            try:
                candidates['date'] = pd.to_datetime(candidates['date'], errors='coerce')
            except Exception:
                pass
            def compute_diff(x):
                if pd.notna(x) and pd.notna(payout_date):
                    try:
                        return abs((payout_date - x).days)
                    except Exception:
                        return np.inf
                else:
                    return np.inf
            candidates['date_diff'] = candidates['date'].apply(compute_diff)
            # drop any non-numeric entries
            candidates['date_diff'] = pd.to_numeric(candidates['date_diff'], errors='coerce').fillna(np.inf)
            candidates = candidates[candidates['date_diff'] <= 3]

        if not candidates.empty:
            candidates['desc_score'] = candidates['description'].apply(desc_score) if 'description' in candidates.columns else 3
            sort_cols = ['amount_diff', 'desc_score']
            if 'date_diff' in candidates.columns:
                sort_cols.append('date_diff')
            candidates = candidates.sort_values(sort_cols)
            bank_idx = candidates.index[0]
            bank_amount = bank_df_copy.loc[bank_idx, 'amount']
            difference = payout_amount - bank_amount
            diff_rounded = round(difference, 2)

            if abs(diff_rounded) <= tolerance:
                status = 'EXACT_MATCH' if diff_rounded == 0 else 'MINOR_VARIANCE'
            else:
                status = 'Mismatch'

            bank_df_copy.loc[bank_idx, 'matched'] = True
            notes = ''
            found = True
            # capture bank description if available
            if 'description' in bank_df_copy.columns:
                bank_desc = bank_df_copy.loc[bank_idx, 'description']
            logger.info(f"Payout {payout_id} matched in bank: {payout_amount} (diff {difference})")
        else:
            bank_amount = np.nan
            difference = payout_amount
            status = 'Missing in Bank'
            notes = ''
            logger.warning(f"Payout {payout_id} not found in bank deposits within tolerance {tolerance}")

        bank_desc = bank_df_copy.loc[bank_idx, 'description'] if ('description' in bank_df_copy.columns and found) else ''
        matching_results.append({
            'payout_id': payout_id,
            'Stripe Amount': payout_amount,
            'Bank Amount': bank_amount,
            'Difference': difference,
            'Status': status,
            'Bank_Row': bank_idx if found and 'bank_idx' in locals() else None,
            'Bank Description': bank_desc,
            'Notes': notes
        })

    # post-processing for partial/combined matches
    unmatched_bank = bank_df_copy[bank_df_copy['matched'] != True].copy()
    unmatched_payouts = [r for r in matching_results if r['Status'] == 'Missing in Bank']

    # try partial matches (one payout -> multiple bank deposits)
    import itertools
    for result in unmatched_payouts:
        pid = result['payout_id']
        amount = result['Stripe Amount']
        candidates = unmatched_bank.copy()
        found_combo = False
        # try combinations of up to 2 deposits
        for k in range(2, 3):
            for combo in itertools.combinations(candidates.index, k):
                total = candidates.loc[list(combo), 'amount'].sum()
                if abs(total - amount) <= tolerance:
                    result['Status'] = 'PARTIAL_MATCH'
                    result['Notes'] = f"Split across {k} bank deposits: {list(combo)}"
                    bank_df_copy.loc[list(combo), 'matched'] = True
                    found_combo = True
                    break
            if found_combo:
                break
    # update unmatched_bank after marking
    unmatched_bank = bank_df_copy[bank_df_copy['matched'] != True].copy()

    # try combined matches (one bank deposit -> multiple payouts)
    remaining_payouts = [r for r in matching_results if r['Status'] == 'Missing in Bank']
    for idx_b, bank_row in unmatched_bank.iterrows():
        bamt = bank_row['amount']
        # look for combos of payouts
        for k in range(2, 3):
            for combo in itertools.combinations(remaining_payouts, k):
                total = sum(r['Stripe Amount'] for r in combo)
                if abs(total - bamt) <= tolerance:
                    # mark those payouts as combined
                    for r in combo:
                        r['Status'] = 'COMBINED_MATCH'
                        r['Notes'] = f"Combined in bank deposit {idx_b}"
                    bank_df_copy.loc[idx_b, 'matched'] = True
                    break
            if bank_df_copy.loc[idx_b, 'matched']:
                break
    # refresh unmatched_bank
    unmatched_bank = bank_df_copy[bank_df_copy['matched'] != True].copy()

    matching_df = pd.DataFrame(matching_results)
    logger.info(
        f"Payout Matching complete: {(matching_df['Status'].isin(['EXACT_MATCH','MINOR_VARIANCE','PARTIAL_MATCH','COMBINED_MATCH'])).sum()} matched, {(matching_df['Status'] == 'Missing in Bank').sum()} missing"
    )

    return matching_df, unmatched_bank


# ============================================================================
# MISSING AND PENDING PAYOUT IDENTIFICATION
# ============================================================================

def identify_missing_payouts(balance_df: pd.DataFrame, payout_df: pd.DataFrame) -> pd.DataFrame:
    """
    Identify transactions in Stripe Balance Report without payout_id or with 
    payout_id not found in Stripe Payout Report.
    Groups by payout_id and sums net amounts.
    Flags as Pending (no payout_id) or Missing (payout_id not in payout report).
    """
    missing_results = []
    payout_ids_in_report = set(payout_df['payout_id'].dropna().unique())
    
    # Find transactions without payout_id
    pending_txns = balance_df[balance_df['payout_id'].isna()].copy()
    if len(pending_txns) > 0:
        pending_net = pending_txns['net'].sum()
        missing_results.append({
            'payout_id': 'PENDING',
            'Net Amount': pending_net,
            'Flag Type': 'Pending Payout',
            'Transaction Count': len(pending_txns)
        })
        logger.warning(f"Found {len(pending_txns)} transactions without payout_id (net: {pending_net})")
    
    # Find payout_ids in balance report but not in payout report
    balance_payout_ids = balance_df[balance_df['payout_id'].notna()]['payout_id'].unique()
    missing_payout_ids = set(balance_payout_ids) - payout_ids_in_report
    
    for payout_id in missing_payout_ids:
        payout_txns = balance_df[balance_df['payout_id'] == payout_id]
        net_amount = payout_txns['net'].sum()
        missing_results.append({
            'payout_id': payout_id,
            'Net Amount': net_amount,
            'Flag Type': 'Missing Payout',
            'Transaction Count': len(payout_txns)
        })
        logger.warning(f"Payout {payout_id} in balance report but not in payout report (net: {net_amount})")
    
    missing_df = pd.DataFrame(missing_results)
    logger.info(f"Missing Payouts identified: {len(missing_df)} entries")
    
    return missing_df


# ============================================================================
# EXCEPTION DETECTION FUNCTIONS
# ============================================================================

def detect_exceptions(
    balance_df: pd.DataFrame,
    payout_df: pd.DataFrame,
    bank_df: pd.DataFrame,
    summary: Dict[str, float],
    matching_df: pd.DataFrame,
    unmatched_bank: pd.DataFrame
) -> pd.DataFrame:
    """
    Detect and flag 15 types of exceptions in reconciliation data.
    Returns DataFrame with exception details for reporting.
    """
    exceptions = []
    
    # Exception 1: Negative Ending Balance
    if summary['Ending Balance'] < 0:
        exceptions.append({
            'Exception_Type': 'Negative Ending Balance',
            'Reference_ID': 'SUMMARY',
            'Amount': summary['Ending Balance'],
            'Explanation': f"Ending balance is negative: {summary['Ending Balance']}"
        })
        logger.error(f"EXCEPTION: Negative Ending Balance: {summary['Ending Balance']}")
    
    # Exception 2: Large Positive Ending Balance (define as > 100,000)
    if summary['Ending Balance'] > 100000:
        exceptions.append({
            'Exception_Type': 'Large Positive Ending Balance',
            'Reference_ID': 'SUMMARY',
            'Amount': summary['Ending Balance'],
            'Explanation': f"Ending balance exceeds 100,000: {summary['Ending Balance']}"
        })
        logger.warning(f"EXCEPTION: Large Positive Ending Balance: {summary['Ending Balance']}")
    
    # Exception 3: Payout not found in bank
    missing_in_bank = matching_df[matching_df['Status'] == 'Missing in Bank']
    for idx, row in missing_in_bank.iterrows():
        exceptions.append({
            'Exception_Type': 'Missing Payout in Bank',
            'Reference_ID': row['payout_id'],
            'Amount': row['Stripe Amount'],
            'Explanation': f"Payout {row['payout_id']} for {row['Stripe Amount']} not found in bank deposits"
        })
        logger.error(f"EXCEPTION: Payout {row['payout_id']} missing in bank")
    
    # Exception 4: Bank deposit not found in payout report
    for idx, row in unmatched_bank.iterrows():
        exceptions.append({
            'Exception_Type': 'Bank Deposit Not in Payout Report',
            'Reference_ID': f"Bank_{row['date'].strftime('%Y%m%d') if pd.notna(row['date']) else 'UNKNOWN'}",
            'Amount': row['amount'],
            'Explanation': f"Bank deposit of {row['amount']} on {row['date']} not found in Stripe payout report"
        })
        logger.warning(f"EXCEPTION: Bank deposit of {row['amount']} not in payout report")
    
    # Exception 5: Duplicate payout_id in payout report
    duplicate_payouts = payout_df[payout_df.duplicated(subset=['payout_id'], keep=False)]
    if len(duplicate_payouts) > 0:
        for payout_id in duplicate_payouts['payout_id'].unique():
            exceptions.append({
                'Exception_Type': 'Duplicate Payout ID',
                'Reference_ID': payout_id,
                'Amount': duplicate_payouts[duplicate_payouts['payout_id'] == payout_id]['amount'].sum(),
                'Explanation': f"Payout {payout_id} appears {len(duplicate_payouts[duplicate_payouts['payout_id'] == payout_id])} times in payout report"
            })
            logger.error(f"EXCEPTION: Duplicate payout_id: {payout_id}")
    
    # Exception 6: Duplicate bank deposits
    duplicate_bank = bank_df[bank_df.duplicated(subset=['date', 'amount'], keep=False)]
    if len(duplicate_bank) > 0:
        for date, amount in duplicate_bank.groupby(['date', 'amount']).size().items():
            exceptions.append({
                'Exception_Type': 'Duplicate Bank Deposit',
                'Reference_ID': f"Bank_{date[0].strftime('%Y%m%d') if pd.notna(date[0]) else 'UNKNOWN'}",
                'Amount': date[1],
                'Explanation': f"Bank deposit of {date[1]} on {date[0]} appears {amount} times"
            })
            logger.warning(f"EXCEPTION: Duplicate bank deposit: {amount} on {date[0]}")
    
    # Exception 7: Mismatch in payout amounts (not exact matches)
    mismatched = matching_df[matching_df['Status'] == 'Mismatch']
    for idx, row in mismatched.iterrows():
        exceptions.append({
            'Exception_Type': 'Payout Amount Mismatch',
            'Reference_ID': row['payout_id'],
            'Amount': row['Difference'],
            'Explanation': f"Payout {row['payout_id']}: Stripe={row['Stripe Amount']}, Bank={row['Bank Amount']}, Difference={row['Difference']}"
        })
        logger.warning(f"EXCEPTION: Payout mismatch: {row['payout_id']}")
    
    # Exception 8: Transactions without payout_id
    pending_count = len(balance_df[balance_df['payout_id'].isna()])
    if pending_count > 0:
        pending_net = balance_df[balance_df['payout_id'].isna()]['net'].sum()
        exceptions.append({
            'Exception_Type': 'Transactions Without Payout ID',
            'Reference_ID': 'PENDING',
            'Amount': pending_net,
            'Explanation': f"{pending_count} transactions without payout_id (net: {pending_net})"
        })
        logger.warning(f"EXCEPTION: {pending_count} transactions without payout_id")
    
    # Exception 9: Currency mismatch between files
    balance_currencies = balance_df['currency'].unique()
    payout_currencies = payout_df['currency'].unique()
    bank_has_currency = len(bank_df) > 0  # Bank statement may not have currency
    
    if len(balance_currencies) > 1 or len(payout_currencies) > 1:
        exceptions.append({
            'Exception_Type': 'Currency Mismatch',
            'Reference_ID': 'MULTI-CURRENCY',
            'Amount': 0,
            'Explanation': f"Multiple currencies detected. Balance: {balance_currencies}, Payout: {payout_currencies}"
        })
        logger.warning(f"EXCEPTION: Currency mismatch detected")

    # Exception: Missing opening balance
    if summary.get('Opening Balance', 0) == 0:
        exceptions.append({
            'Exception_Type': 'Missing Opening Balance',
            'Reference_ID': 'SUMMARY',
            'Amount': 0,
            'Explanation': 'Opening balance is zero; verify prior period adjustment'
        })
        logger.warning("EXCEPTION: Opening balance appears to be zero")

    # Exception: Partial or combined payouts
    for idx, row in matching_df.iterrows():
        status = row.get('Status', '')
        if status == 'PARTIAL_MATCH':
            exceptions.append({
                'Exception_Type': 'Partial Payout',
                'Reference_ID': row.get('payout_id'),
                'Amount': row.get('Stripe Amount'),
                'Explanation': 'Payout split across multiple bank deposits',
                'Severity': 'Medium'
            })
            logger.warning(f"EXCEPTION: Partial payout {row.get('payout_id')} detected")
        if status == 'COMBINED_MATCH':
            exceptions.append({
                'Exception_Type': 'Combined Payout',
                'Reference_ID': row.get('payout_id'),
                'Amount': row.get('Stripe Amount'),
                'Explanation': 'Payout combined with others in one bank deposit',
                'Severity': 'Medium'
            })
            logger.warning(f"EXCEPTION: Combined payout {row.get('payout_id')} detected")
        # description mismatch check
        desc = str(row.get('Bank Description', '')).lower()
        pid = str(row.get('payout_id', '')).lower()
        if status in ['EXACT_MATCH', 'MINOR_VARIANCE'] and desc:
            if pid and pid not in desc and 'stripe' not in desc:
                exceptions.append({
                    'Exception_Type': 'Bank Description Mismatch',
                    'Reference_ID': row.get('payout_id'),
                    'Amount': row.get('Stripe Amount'),
                    'Explanation': f"Bank description '{row.get('Bank Description')}' does not reference payout {row.get('payout_id')} or stripe"
                })
                logger.warning(f"EXCEPTION: Bank description mismatch for payout {row.get('payout_id')}" )
    # Exception 11: Empty payout report
    if len(payout_df) == 0:
        exceptions.append({
            'Exception_Type': 'Empty Payout Report',
            'Reference_ID': 'EMPTY',
            'Amount': 0,
            'Explanation': 'Stripe Payout Report contains no records'
        })
        logger.error(f"EXCEPTION: Empty payout report")
    
    # Exception 12: Refund assigned to later payout (check dates)
    # FUTURE ENHANCEMENT: Add date-based validation for refunds
    
    # Exception 13: Dispute without payout (transactions without payout_id)
    disputes_without_payout = balance_df[
        (balance_df['type'].str.lower() == 'dispute') & 
        (balance_df['payout_id'].isna())
    ]
    if len(disputes_without_payout) > 0:
        exceptions.append({
            'Exception_Type': 'Dispute Without Payout ID',
            'Reference_ID': 'DISPUTES',
            'Amount': disputes_without_payout['net'].sum(),
            'Explanation': f"{len(disputes_without_payout)} disputes without payout_id"
        })
        logger.warning(f"EXCEPTION: {len(disputes_without_payout)} disputes without payout")
    
    exceptions_df = pd.DataFrame(exceptions)
    logger.info(f"Exception Detection complete: {len(exceptions_df)} exceptions found")
    
    return exceptions_df


# ============================================================================
# EXCEL OUTPUT GENERATION
# ============================================================================

def create_excel_output(
    output_file: str,
    summary: Dict[str, float],
    matching_df: pd.DataFrame,
    missing_df: pd.DataFrame,
    unmatched_bank: pd.DataFrame,
    exceptions_df: pd.DataFrame
) -> None:
    """
    Create Excel workbook with 5 sheets: Stripe_Monthly_Summary, Payout_Matching,
    Missing_Payouts, Bank_Unmatched, and Exception_Report.
    Applies formatting and styling to make output professional and readable.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define cell styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========================================================================
    # Sheet 1: Stripe Monthly Summary
    # ========================================================================
    ws_summary = wb.create_sheet('Stripe_Monthly_Summary', 0)
    ws_summary['A1'] = 'Metric'
    ws_summary['B1'] = 'Amount'
    
    # Apply header formatting
    for col in ['A1', 'B1']:
        cell = ws_summary[col]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add summary metrics
    row = 2
    for metric, value in summary.items():
        ws_summary[f'A{row}'] = metric
        ws_summary[f'B{row}'] = value if isinstance(value, str) else round(value, 2)
        
        # Apply cell formatting
        for col in ['A', 'B']:
            cell = ws_summary[f'{col}{row}']
            cell.border = border
            if col == 'B' and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        
        row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    logger.info("Created Stripe_Monthly_Summary sheet")
    
    # ========================================================================
    # Sheet 2: Payout Matching
    # ========================================================================
    ws_matching = wb.create_sheet('Payout_Matching', 1)
    # dynamic headers from dataframe
    headers = list(matching_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_matching.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add matching data
    for row_idx, row_data in enumerate(matching_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_matching.cell(row=row_idx, column=col_idx)
            cell.value = value if not pd.isna(value) else ''
            cell.border = border
            
            # Format numeric columns (amounts and differences)
            if col_idx in [2, 3, 4] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
            # wrap text in notes or description columns
            if headers[col_idx-1].lower() in ['notes', 'bank description']:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths generically
    for i, header in enumerate(headers, 1):
        width = 20
        if 'amount' in header.lower() or 'difference' in header.lower():
            width = 18
        if header.lower() in ['notes', 'bank description']:
            width = 40
        ws_matching.column_dimensions[chr(64 + i)].width = width
    
    logger.info("Created Payout_Matching sheet")
    
    # ========================================================================
    # Sheet 3: Missing Payouts
    # ========================================================================
    ws_missing = wb.create_sheet('Missing_Payouts', 2)
    headers = ['payout_id', 'Net Amount', 'Flag Type', 'Transaction Count']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_missing.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add missing payout data
    for row_idx, row_data in enumerate(missing_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_missing.cell(row=row_idx, column=col_idx)
            cell.value = value if not pd.isna(value) else ''
            cell.border = border
            
            # Format numeric columns
            if col_idx in [2, 4] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
    
    # Adjust column widths
    ws_missing.column_dimensions['A'].width = 20
    ws_missing.column_dimensions['B'].width = 18
    ws_missing.column_dimensions['C'].width = 20
    ws_missing.column_dimensions['D'].width = 20
    
    logger.info("Created Missing_Payouts sheet")
    
    # ========================================================================
    # Sheet 4: Bank Unmatched
    # ========================================================================
    ws_bank = wb.create_sheet('Bank_Unmatched', 3)
    headers = ['date', 'description', 'amount', 'Reason']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_bank.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add unmatched bank data with reason
    for row_idx, (idx, row_data) in enumerate(unmatched_bank.iterrows(), 2):
        ws_bank.cell(row=row_idx, column=1).value = row_data['date']
        ws_bank.cell(row=row_idx, column=2).value = row_data['description']
        ws_bank.cell(row=row_idx, column=3).value = row_data['amount']
        ws_bank.cell(row=row_idx, column=4).value = 'Not in Stripe Payout Report'
        
        for col_idx in range(1, 5):
            cell = ws_bank.cell(row=row_idx, column=col_idx)
            cell.border = border
            if col_idx == 3 and isinstance(row_data['amount'], (int, float)):
                cell.number_format = '#,##0.00'
            elif col_idx == 1:
                cell.number_format = 'YYYY-MM-DD'
    
    # Adjust column widths
    ws_bank.column_dimensions['A'].width = 15
    ws_bank.column_dimensions['B'].width = 35
    ws_bank.column_dimensions['C'].width = 18
    ws_bank.column_dimensions['D'].width = 35
    
    logger.info("Created Bank_Unmatched sheet")
    
    # ========================================================================
    # Sheet 5: Exception Report
    # ========================================================================
    ws_exceptions = wb.create_sheet('Exception_Report', 4)
    headers = list(exceptions_df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_exceptions.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add exception data
    for row_idx, row_data in enumerate(exceptions_df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_exceptions.cell(row=row_idx, column=col_idx)
            cell.value = value if not pd.isna(value) else ''
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Format numeric columns automatically
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
    
    # Adjust column widths heuristically
    for i, header in enumerate(headers, 1):
        width = 20
        if 'amount' in header.lower():
            width = 18
        if 'explanation' in header.lower() or 'notes' in header.lower():
            width = 50
        ws_exceptions.column_dimensions[chr(64 + i)].width = width
    
    logger.info("Created Exception_Report sheet")
    
    # Save workbook
    wb.save(output_file)
    logger.info(f"Excel output created successfully: {output_file}")


# ============================================================================
# MAIN EXECUTION FUNCTION
# ============================================================================

def main(
    stripe_balance_file: str,
    stripe_payout_file: str,
    bank_statement_file: str,
    output_file: str = 'reconciliation_output.xlsx',
    opening_balance: float = 0.0,
    tolerance: float = 5.0
) -> None:
    """
    Main execution function that orchestrates the entire reconciliation process.
    Loads input files, performs all calculations and exception detection,
    and generates Excel output report.  Accepts an opening balance and
    tolerance for matching and summary status checks.
    """
    try:
        logger.info("="*80)
        logger.info("Starting Stripe Reconciliation Process")
        logger.info(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("="*80)
        
        # ====================================================================
        # Step 1: Load Input Files
        # ====================================================================
        logger.info("\n[STEP 1] Loading input files...")
        balance_df = load_stripe_balance_report(stripe_balance_file)
        payout_df = load_stripe_payout_report(stripe_payout_file)
        bank_df = load_bank_statement(bank_statement_file)
        
        # Add helper column for tracking matched bank deposits
        bank_df['matched'] = False
        
        # ====================================================================
        # Step 2: Calculate Monthly Summary
        # ====================================================================
        logger.info("\n[STEP 2] Calculating Stripe monthly summary...")
        summary = calculate_stripe_monthly_summary(
            balance_df, payout_df, bank_df, opening_balance, tolerance
        )
        
        # ====================================================================
        # Step 3: Match Payouts to Bank Deposits
        # ====================================================================
        logger.info("\n[STEP 3] Matching payouts to bank deposits...")
        matching_df, unmatched_bank = match_stripe_payouts_to_bank(
            payout_df, bank_df, tolerance
        )
        
        # ====================================================================
        # Step 4: Identify Missing Payouts
        # ====================================================================
        logger.info("\n[STEP 4] Identifying missing and pending payouts...")
        missing_df = identify_missing_payouts(balance_df, payout_df)
        
        # ====================================================================
        # Step 5: Detect Exceptions
        # ====================================================================
        logger.info("\n[STEP 5] Detecting exceptions...")
        exceptions_df = detect_exceptions(balance_df, payout_df, bank_df, summary, matching_df, unmatched_bank)
        
        # Update reconciliation status based on exceptions
        if len(exceptions_df) == 0:
            summary['Reconciliation Status'] = 'Reconciled'
        else:
            summary['Reconciliation Status'] = 'Exception'
        
        # ====================================================================
        # Step 6: Generate Excel Output
        # ====================================================================
        logger.info("\n[STEP 6] Generating Excel output...")
        create_excel_output(output_file, summary, matching_df, missing_df, unmatched_bank, exceptions_df)
        
        # ====================================================================
        # Final Summary
        # ====================================================================
        logger.info("\n" + "="*80)
        logger.info("RECONCILIATION COMPLETE")
        logger.info("="*80)
        logger.info(f"Output file: {output_file}")
        logger.info(f"Exceptions found: {len(exceptions_df)}")
        logger.info(f"Reconciliation Status: {summary['Reconciliation Status']}")
        logger.info("="*80 + "\n")
        
    except FileNotFoundError as e:
        logger.error(f"File not found: {str(e)}")
        raise
    except ValueError as e:
        logger.error(f"Validation error: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        raise


# ============================================================================
# STRIPE RECONCILIATOR CLASS FOR WEB APPLICATION
# ============================================================================

class StripeReconciliator:
    """
    Web-application-friendly wrapper around reconciliation logic.
    Handles in-memory file processing and report generation.
    """
    
    def __init__(self):
        """Initialize reconciliator with empty dataframes."""
        self.balance_df = None
        self.payout_df = None
        self.bank_df = None
        self.exceptions = []
        self.summary = {}
        self.matching_df = None
        self.missing_df = None
        self.unmatched_bank = None
        self.exceptions_df = None
        
    def load_csv(self, file_content: bytes, file_type: str) -> bool:
        """
        Load a CSV file from bytes content.
        
        Args:
            file_content: CSV file content as bytes
            file_type: Type of file ('balance', 'payout', or 'bank')
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Convert bytes to string
            csv_string = file_content.decode('utf-8')
            
            # Parse CSV
            if file_type == 'balance':
                try:
                    self.balance_df = self._load_balance_csv(csv_string)
                    logger.info(f"Balance file loaded successfully: {len(self.balance_df)} rows")
                    return True
                except Exception as e:
                    msg = f"Failed to parse balance file: {str(e)}"
                    logger.error(msg)
                    self.exceptions.append(msg)
                    return False
            elif file_type == 'payout':
                try:
                    self.payout_df = self._load_payout_csv(csv_string)
                    logger.info(f"Payout file loaded successfully: {len(self.payout_df)} rows")
                    return True
                except Exception as e:
                    msg = f"Failed to parse payout file: {str(e)}"
                    logger.error(msg)
                    self.exceptions.append(msg)
                    return False
            elif file_type == 'bank':
                try:
                    self.bank_df = self._load_bank_csv(csv_string)
                    logger.info(f"Bank statement loaded successfully: {len(self.bank_df)} rows")
                    return True
                except Exception as e:
                    msg = f"Failed to parse bank statement: {str(e)}"
                    logger.error(msg)
                    self.exceptions.append(msg)
                    return False
            else:
                msg = f"Unknown file type: {file_type}"
                logger.error(msg)
                self.exceptions.append(msg)
                return False
                
        except UnicodeDecodeError as e:
            msg = f"File encoding error: {str(e)}. Please ensure the file is UTF-8 encoded."
            logger.error(msg)
            self.exceptions.append(msg)
            return False
        except Exception as e:
            msg = f"Unexpected error loading {file_type} file: {str(e)}"
            logger.error(msg)
            self.exceptions.append(msg)
            return False
    
    def _load_balance_csv(self, csv_string: str) -> pd.DataFrame:
        """Load balance CSV from string."""
        df = pd.read_csv(StringIO(csv_string))
        
        logger.info(f"Balance CSV columns: {list(df.columns)}")
        
        # Balance sheet columns - match to both sample data and Stripe API formats
        balance_columns = {
            'created': ['date', 'Date', 'Transaction Date', 'created', 'Timestamp'],
            'payout_id': ['payout id', 'Payout ID', 'payout_id', 'id', 'ID'],
            'type': ['type', 'Type', 'transaction_type', 'Transaction Type'],
            'amount': ['charge amount', 'Charge Amount', 'amount', 'Amount', 'Transaction Amount'],
            'fee': ['fee amount', 'Fee Amount', 'fee', 'Fee'],
            'net': ['balance', 'Balance', 'net_amount', 'Net Amount', 'net']
        }
        
        try:
            mapping = get_column_mapping(
                df.columns.tolist(),
                balance_columns,
                ['created', 'payout_id'],
                optional_columns={'type': 'charge', 'amount': 0, 'fee': 0, 'net': 0}
            )
            
            df = rename_columns_based_on_mapping(df, mapping)
            
        except ValueError as e:
            logger.warning(f"Custom column mapping failed: {str(e)}. Using defaults.")
            # Try a simpler approach - just rename what we can find
            rename_map = {}
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'date' in col_lower or 'timestamp' in col_lower:
                    rename_map[col] = 'created'
                elif 'payout' in col_lower and 'id' in col_lower:
                    rename_map[col] = 'payout_id'
                elif 'charge' in col_lower:
                    rename_map[col] = 'amount'
                elif 'fee' in col_lower:
                    rename_map[col] = 'fee'
                elif 'balance' in col_lower:
                    rename_map[col] = 'net'
            
            if rename_map:
                df = df.rename(columns=rename_map)
            else:
                raise ValueError(f"Could not map any columns. Available: {list(df.columns)}")
        
        # Ensure required columns exist with defaults
        for col in ['created', 'payout_id', 'type', 'amount', 'fee', 'net', 'currency']:
            if col not in df.columns:
                if col == 'type':
                    df[col] = 'charge'
                elif col == 'currency':
                    df[col] = 'USD'
                else:
                    df[col] = 0 if col in ['amount', 'fee', 'net'] else None
            
        # Convert datetime and numeric columns
        df['created'] = pd.to_datetime(df['created'], errors='coerce')
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
        df['fee'] = pd.to_numeric(df['fee'], errors='coerce').fillna(0)
        df['net'] = pd.to_numeric(df['net'], errors='coerce').fillna(0)
        
        logger.info(f"Balance CSV processed. Final columns: {list(df.columns)}")
        return df
    
    def _load_payout_csv(self, csv_string: str) -> pd.DataFrame:
        """Load payout CSV from string."""
        df = pd.read_csv(StringIO(csv_string))
        
        logger.info(f"Payout CSV columns: {list(df.columns)}")
        
        # Payout columns - flexible mapping for variations
        payout_columns = {
            'payout_id': ['payout id', 'Payout ID', 'payout_id', 'id', 'ID'],
            'created': ['payout date', 'Payout Date', 'created', 'date', 'Date', 'arrival date', 'Arrival Date'],
            'amount': ['payout amount', 'Payout Amount', 'amount', 'Amount'],
            'currency': ['currency', 'Currency'],
            'status': ['status', 'Status'],
            'type': ['type', 'Type'],
            'arrival_date': ['arrival date', 'Arrival Date', 'arrival_date', 'payout date', 'Payout Date']
        }
        
        try:
            mapping = get_column_mapping(
                df.columns.tolist(),
                payout_columns,
                ['payout_id', 'created', 'amount', 'status'],
                optional_columns={'currency': 'USD', 'type': 'payout', 'arrival_date': 'created'}
            )
            
            df = rename_columns_based_on_mapping(df, mapping)
            
        except ValueError as e:
            logger.warning(f"Custom column mapping failed: {str(e)}. Using defaults.")
            # Try a simpler approach - look for arrival_date as fallback for created
            rename_map = {}
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'payout' in col_lower and 'id' in col_lower:
                    rename_map[col] = 'payout_id'
                elif 'payout date' in col_lower or ('date' in col_lower and 'arrival' not in col_lower):
                    if 'created' not in rename_map.values():
                        rename_map[col] = 'created'
                elif 'arrival' in col_lower:
                    if 'arrival_date' not in rename_map.values():
                        rename_map[col] = 'arrival_date'
                elif 'amount' in col_lower:
                    rename_map[col] = 'amount'
                elif 'status' in col_lower:
                    rename_map[col] = 'status'
                elif 'currency' in col_lower:
                    rename_map[col] = 'currency'
            
            if 'payout_id' not in rename_map.values():
                raise ValueError(f"Could not find payout ID column. Available: {list(df.columns)}")
            if 'amount' not in rename_map.values():
                raise ValueError(f"Could not find amount column. Available: {list(df.columns)}")
            
            if rename_map:
                df = df.rename(columns=rename_map)
        
        # If 'created' is missing but 'arrival_date' exists, use arrival_date as created
        if 'created' not in df.columns and 'arrival_date' in df.columns:
            df['created'] = df['arrival_date']
            logger.info("Using 'arrival_date' as 'created' date")
        
        # Ensure required columns exist
        for col in ['payout_id', 'created', 'amount', 'status', 'currency', 'type', 'arrival_date']:
            if col not in df.columns:
                if col == 'currency':
                    df[col] = 'USD'
                elif col == 'type':
                    df[col] = 'payout'
                elif col == 'status':
                    df[col] = 'paid'
                elif col == 'arrival_date':
                    df[col] = df.get('created', pd.NaT)
                else:
                    df[col] = None
            
        # Convert datetime and numeric columns
        df['created'] = pd.to_datetime(df['created'], errors='coerce')
        if 'arrival_date' in df.columns:
            df['arrival_date'] = pd.to_datetime(df['arrival_date'], errors='coerce')
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
        
        logger.info(f"Payout CSV processed. Final columns: {list(df.columns)}")
        return df
    
    def _load_bank_csv(self, csv_string: str) -> pd.DataFrame:
        """Load bank statement CSV from string."""
        df = pd.read_csv(StringIO(csv_string))
        
        logger.info(f"Bank CSV columns: {list(df.columns)}")
        
        # Bank statement columns - prioritize 'Credit' as the deposit amount
        bank_columns = {
            'date': ['date', 'Date', 'Transaction Date', 'transaction_date', 'Timestamp'],
            'description': ['description', 'Description', 'Memo', 'memo', 'Reference'],
            'debit': ['debit', 'Debit'],
            'credit': ['credit', 'Credit', 'Deposit', 'amount', 'Amount'],
            'balance': ['balance', 'Balance']
        }
        
        try:
            # For bank statement, we need at least date and a credit/amount column
            mapping = get_column_mapping(
                df.columns.tolist(),
                bank_columns,
                ['date', 'credit'],
                optional_columns={'description': '', 'debit': 0, 'balance': 0}
            )
            
            df = rename_columns_based_on_mapping(df, mapping)
            
        except ValueError as e:
            logger.warning(f"Custom column mapping failed: {str(e)}. Using defaults.")
            # Try a simpler approach - find required columns
            rename_map = {}
            
            # Find date column
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'date' in col_lower or 'timestamp' in col_lower:
                    if 'date' not in rename_map.values():
                        rename_map[col] = 'date'
                        break
            
            # Find amount/credit column (prefer 'credit' for deposits)
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'credit' in col_lower or 'deposit' in col_lower:
                    rename_map[col] = 'credit'
                    break
            
            # If no credit, look for amount
            if 'credit' not in rename_map.values():
                for col in df.columns:
                    col_lower = col.lower().strip()
                    if 'amount' in col_lower:
                        rename_map[col] = 'credit'
                        break
            
            # Find other optional columns
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'description' in col_lower or 'memo' in col_lower:
                    rename_map[col] = 'description'
                elif 'debit' in col_lower:
                    rename_map[col] = 'debit'
                elif 'balance' in col_lower:
                    rename_map[col] = 'balance'
            
            if 'date' not in rename_map.values():
                raise ValueError(f"Could not find date column. Available: {list(df.columns)}")
            if 'credit' not in rename_map.values():
                raise ValueError(f"Could not find deposit/credit/amount column. Available: {list(df.columns)}")
            
            if rename_map:
                df = df.rename(columns=rename_map)
        
        # Rename 'credit' to 'amount' for consistency with reconciliation functions
        if 'credit' in df.columns and 'amount' not in df.columns:
            df = df.rename(columns={'credit': 'amount'})
        
        # Ensure required columns exist
        for col in ['date', 'amount', 'description', 'debit', 'balance']:
            if col not in df.columns:
                if col == 'description':
                    df[col] = ''
                else:
                    df[col] = 0
            
        # Convert columns
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['debit'] = pd.to_numeric(df['debit'], errors='coerce').fillna(0)
        df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
        df['balance'] = pd.to_numeric(df['balance'], errors='coerce').fillna(0)
        
        df['matched'] = False
        
        logger.info(f"Bank CSV processed. Final columns: {list(df.columns)}")
        return df
    
    def process_files(self, opening_balance: float = 0.0, tolerance: float = 5.0) -> bool:
        """
        Process loaded files and perform reconciliation.
        
        Args:
            opening_balance: Opening balance for reconciliation
            tolerance: Tolerance for amount matching
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Validate all files are loaded
            if not all([self.balance_df is not None, self.payout_df is not None, self.bank_df is not None]):
                msg = "Not all files loaded. Please upload balance, payout, and bank statement files."
                logger.error(msg)
                self.exceptions.append(msg)
                return False
            
            logger.info(f"Processing reconciliation with {len(self.balance_df)} balance rows, {len(self.payout_df)} payout rows, {len(self.bank_df)} bank rows")
            
            try:
                # Calculate monthly summary
                logger.info("Step 1: Calculating monthly summary...")
                self.summary = calculate_stripe_monthly_summary(
                    self.balance_df, self.payout_df, self.bank_df, opening_balance, tolerance
                )
                logger.info(f"Summary calculated: {len(self.summary)} metrics")
            except Exception as e:
                msg = f"Error calculating summary: {str(e)}"
                logger.error(msg, exc_info=True)
                self.exceptions.append(msg)
                return False
            
            try:
                # Match payouts to bank
                logger.info("Step 2: Matching payouts to bank deposits...")
                self.matching_df, self.unmatched_bank = match_stripe_payouts_to_bank(
                    self.payout_df, self.bank_df, tolerance
                )
                logger.info(f"Payout matching complete: {len(self.matching_df)} payouts, {len(self.unmatched_bank)} unmatched bank")
            except Exception as e:
                msg = f"Error matching payouts: {str(e)}"
                logger.error(msg, exc_info=True)
                self.exceptions.append(msg)
                return False
            
            try:
                # Identify missing payouts
                logger.info("Step 3: Identifying missing payouts...")
                self.missing_df = identify_missing_payouts(self.balance_df, self.payout_df)
                logger.info(f"Missing payouts identified: {len(self.missing_df) if self.missing_df is not None else 0}")
            except Exception as e:
                msg = f"Error identifying missing payouts: {str(e)}"
                logger.error(msg, exc_info=True)
                self.exceptions.append(msg)
                return False
            
            try:
                # Detect exceptions
                logger.info("Step 4: Detecting exceptions...")
                self.exceptions_df = detect_exceptions(
                    self.balance_df, self.payout_df, self.bank_df, 
                    self.summary, self.matching_df, self.unmatched_bank
                )
                logger.info(f"Exception detection complete: {len(self.exceptions_df) if self.exceptions_df is not None else 0} exceptions")
            except Exception as e:
                msg = f"Error detecting exceptions: {str(e)}"
                logger.error(msg, exc_info=True)
                self.exceptions.append(msg)
                return False
            
            # Update reconciliation status
            if self.exceptions_df is None or len(self.exceptions_df) == 0:
                self.summary['Reconciliation Status'] = 'Reconciled'
            else:
                self.summary['Reconciliation Status'] = 'Exception'
            
            # Store exception details
            if self.exceptions_df is not None and len(self.exceptions_df) > 0:
                self.exceptions.extend([
                    f"{row.get('Exception_Type', 'Unknown')}: {row.get('Explanation', '')}" 
                    for idx, row in self.exceptions_df.iterrows()
                ])
            
            logger.info(f"Processing complete. Reconciliation Status: {self.summary.get('Reconciliation Status', 'Unknown')}")
            return True
            
        except Exception as e:
            msg = f"Unexpected error processing files: {str(e)}"
            logger.error(msg, exc_info=True)
            self.exceptions.append(msg)
            return False
    
    def generate_report(self) -> BytesIO:
        """
        Generate Excel report from processed data.
        
        Returns:
            BytesIO object containing the Excel file, or None if processing failed
        """
        try:
            if self.summary is None or self.matching_df is None:
                msg = "Files have not been processed yet. Please process the reconciliation first."
                logger.error(msg)
                self.exceptions.append(msg)
                return None
            
            logger.info("Generating Excel report...")
            
            # Create Excel workbook
            wb = Workbook()
            wb.remove(wb.active)
            
            # Define styles
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sheet 1: Summary
            try:
                ws_summary = wb.create_sheet('Summary', 0)
                ws_summary['A1'] = 'Metric'
                ws_summary['B1'] = 'Amount'
                
                for col in ['A1', 'B1']:
                    cell = ws_summary[col]
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                row = 2
                for metric, value in self.summary.items():
                    ws_summary[f'A{row}'] = metric
                    try:
                        ws_summary[f'B{row}'] = value if isinstance(value, str) else round(float(value), 2)
                    except (ValueError, TypeError):
                        ws_summary[f'B{row}'] = str(value)
                    
                    for col in ['A', 'B']:
                        cell = ws_summary[f'{col}{row}']
                        cell.border = border
                        if col == 'B' and isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                    row += 1
                
                ws_summary.column_dimensions['A'].width = 30
                ws_summary.column_dimensions['B'].width = 20
                logger.info("Created Summary sheet")
            except Exception as e:
                logger.error(f"Error creating summary sheet: {str(e)}", exc_info=True)
                raise
            
            # Sheet 2: Payout Matching
            try:
                ws_matching = wb.create_sheet('Payout_Matching', 1)
                headers = list(self.matching_df.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_matching.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                for row_idx, row_data in enumerate(self.matching_df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws_matching.cell(row=row_idx, column=col_idx)
                        try:
                            cell.value = value
                        except:
                            cell.value = str(value)
                        cell.border = border
                
                for col_idx in range(1, len(headers) + 1):
                    ws_matching.column_dimensions[ws_matching.cell(row=1, column=col_idx).column_letter].width = 15
                logger.info("Created Payout_Matching sheet")
            except Exception as e:
                logger.error(f"Error creating payout matching sheet: {str(e)}", exc_info=True)
                raise
            
            # Sheet 3: Missing Payouts
            try:
                if self.missing_df is not None and len(self.missing_df) > 0:
                    ws_missing = wb.create_sheet('Missing_Payouts', 2)
                    headers = list(self.missing_df.columns)
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_missing.cell(row=1, column=col_idx)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    for row_idx, row_data in enumerate(self.missing_df.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_missing.cell(row=row_idx, column=col_idx)
                            try:
                                cell.value = value
                            except:
                                cell.value = str(value)
                            cell.border = border
                    
                    for col_idx in range(1, len(headers) + 1):
                        ws_missing.column_dimensions[ws_missing.cell(row=1, column=col_idx).column_letter].width = 15
                    logger.info("Created Missing_Payouts sheet")
            except Exception as e:
                logger.error(f"Error creating missing payouts sheet: {str(e)}", exc_info=True)
                raise
            
            # Sheet 4: Unmatched Bank
            try:
                if self.unmatched_bank is not None and len(self.unmatched_bank) > 0:
                    ws_unmatched = wb.create_sheet('Unmatched_Bank', 3)
                    headers = list(self.unmatched_bank.columns)
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_unmatched.cell(row=1, column=col_idx)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    for row_idx, row_data in enumerate(self.unmatched_bank.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_unmatched.cell(row=row_idx, column=col_idx)
                            try:
                                cell.value = value
                            except:
                                cell.value = str(value)
                            cell.border = border
                    
                    for col_idx in range(1, len(headers) + 1):
                        ws_unmatched.column_dimensions[ws_unmatched.cell(row=1, column=col_idx).column_letter].width = 15
                    logger.info("Created Unmatched_Bank sheet")
            except Exception as e:
                logger.error(f"Error creating unmatched bank sheet: {str(e)}", exc_info=True)
                raise
            
            # Sheet 5: Exceptions
            try:
                if self.exceptions_df is not None and len(self.exceptions_df) > 0:
                    ws_exceptions = wb.create_sheet('Exceptions', 4)
                    headers = list(self.exceptions_df.columns)
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws_exceptions.cell(row=1, column=col_idx)
                        cell.value = header
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    for row_idx, row_data in enumerate(self.exceptions_df.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws_exceptions.cell(row=row_idx, column=col_idx)
                            try:
                                cell.value = value
                            except:
                                cell.value = str(value)
                            cell.border = border
                    
                    for col_idx in range(1, len(headers) + 1):
                        ws_exceptions.column_dimensions[ws_exceptions.cell(row=1, column=col_idx).column_letter].width = 15
                    logger.info("Created Exceptions sheet")
            except Exception as e:
                logger.error(f"Error creating exceptions sheet: {str(e)}", exc_info=True)
                raise
            
            # Save to BytesIO
            try:
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                logger.info("Report generated successfully")
                return output
            except Exception as e:
                msg = f"Error saving workbook: {str(e)}"
                logger.error(msg, exc_info=True)
                self.exceptions.append(msg)
                return None
            
        except Exception as e:
            msg = f"Error generating report: {str(e)}"
            logger.error(msg, exc_info=True)
            self.exceptions.append(msg)
            return None
